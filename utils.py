import os
import csv
import io
from datetime import datetime, timedelta
import random
from database.models import db, User, SensorData, Predictions, PumpLogs, Weather
from weather import get_weather_data
from ml.predict import predictor

# Excel export using openpyxl
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# PDF export using reportlab
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# Create default admin if no users exist
def init_db_defaults():
    db.create_all()
    # Check if admin exists
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        from werkzeug.security import generate_password_hash
        admin = User(
            username='admin',
            email='admin@irrigation.com',
            password_hash=generate_password_hash('admin123'),
            is_admin=True
        )
        db.session.add(admin)
        db.session.commit()
        print("Default admin user created (admin/admin123).")
        
    # Check if a regular test user exists
    test_user = User.query.filter_by(username='farmer').first()
    if not test_user:
        from werkzeug.security import generate_password_hash
        test_user = User(
            username='farmer',
            email='farmer@field.com',
            password_hash=generate_password_hash('farmer123'),
            is_admin=False
        )
        db.session.add(test_user)
        db.session.commit()
        print("Default farmer user created (farmer/farmer123).")

# Simulate sensor reading
def generate_simulated_reading():
    """
    Generates realistic sensor readings.
    """
    last_reading = SensorData.query.order_by(SensorData.timestamp.desc()).first()
    
    if last_reading:
        # Base on last values with small random variations
        moisture = max(10.0, min(95.0, last_reading.moisture + random.uniform(-2.5, 2.5)))
        temperature = max(15.0, min(45.0, last_reading.temperature + random.uniform(-0.5, 0.5)))
        humidity = max(20.0, min(95.0, last_reading.humidity + random.uniform(-1.0, 1.0)))
    else:
        # Initial defaults
        moisture = 35.0
        temperature = 28.5
        humidity = 60.0
        
    # Rain sensor: check weather condition
    weather = Weather.query.order_by(Weather.timestamp.desc()).first()
    if weather and weather.condition.lower() == 'rain':
        rain_val = random.uniform(10.0, 30.0) # Low analog value = wet/rain detected
        moisture = min(95.0, moisture + random.uniform(3.0, 8.0))
    else:
        rain_val = random.uniform(90.0, 100.0) # High analog value = dry/no rain
        
    return {
        'moisture': round(moisture, 2),
        'temperature': round(temperature, 2),
        'humidity': round(humidity, 2),
        'rain_value': round(rain_val, 2)
    }

# Process sensor logs and make predictions
def process_new_sensor_reading(data, system_state):
    """
    Saves new sensor data, runs prediction, and controls pump if AUTO mode.
    system_state: dict containing config, mode, current_pump_status, crop, etc.
    """
    # 1. Save sensor reading
    sensor_log = SensorData(
        moisture=data['moisture'],
        temperature=data['temperature'],
        humidity=data['humidity'],
        rain_value=data['rain_value']
    )
    db.session.add(sensor_log)
    
    # Is it raining? (Let's define analog rain sensor < 50.0 as rain detected)
    is_raining = data['rain_value'] < 50.0
    
    # 2. Query/Update weather cache (every 30 mins or if city changed)
    last_weather = Weather.query.order_by(Weather.timestamp.desc()).first()
    city = system_state.get('weather_city') or 'Delhi'
    if not last_weather or last_weather.city.lower() != city.lower() or (datetime.utcnow() - last_weather.timestamp) > timedelta(minutes=30):
        w_data = get_weather_data(city)
        weather_log = Weather(
            city=w_data['city'],
            temperature=w_data['temperature'],
            humidity=w_data['humidity'],
            wind_speed=w_data['wind_speed'],
            rain_prob=w_data['rain_prob'],
            condition=w_data['condition']
        )
        db.session.add(weather_log)
        db.session.commit()
        last_weather = weather_log
        
    # Weather forecast override
    weather_forecast_rain = last_weather.rain_prob > 60.0 or last_weather.condition.lower() == 'rain'
    
    # 3. Trigger Machine Learning Prediction
    # Crop mapping: crop type string from system state
    crop = system_state.get('crop', 'Wheat')
    rain_input = 1 if (is_raining or weather_forecast_rain) else 0
    
    pred_str, confidence = predictor.predict(
        moisture=data['moisture'],
        temperature=data['temperature'],
        humidity=data['humidity'],
        rain=rain_input,
        crop_type=crop
    )
    
    pred_log = Predictions(
        moisture=data['moisture'],
        temperature=data['temperature'],
        humidity=data['humidity'],
        rain=bool(rain_input),
        crop_type=crop,
        prediction=pred_str,
        confidence=confidence
    )
    db.session.add(pred_log)
    
    # 4. Automation control
    if system_state.get('mode') == 'AUTO':
        new_status = 'ON' if pred_str == 'Pump ON' else 'OFF'
        old_status = system_state.get('pump_status', 'OFF')
        
        # Override to OFF if rain is actively detected or heavy rain predicted
        if rain_input == 1:
            new_status = 'OFF'
            
        if new_status != old_status:
            system_state['pump_status'] = new_status
            log = PumpLogs(action=f"Pump {new_status}", source="AUTO")
            db.session.add(log)
            print(f"[AUTO] Pump status changed to {new_status} based on ML predictions.")
            
    db.session.commit()

# Export functions
def export_csv(table_name):
    output = io.StringIO()
    writer = csv.writer(output)
    
    if table_name == 'sensor_data':
        writer.writerow(['ID', 'Moisture (%)', 'Temperature (°C)', 'Humidity (%)', 'Rain Value', 'Timestamp'])
        records = SensorData.query.order_by(SensorData.timestamp.desc()).limit(1000).all()
        for r in records:
            writer.writerow([r.id, r.moisture, r.temperature, r.humidity, r.rain_value, r.timestamp])
            
    elif table_name == 'pump_logs':
        writer.writerow(['ID', 'Action', 'Source', 'Timestamp'])
        records = PumpLogs.query.order_by(PumpLogs.timestamp.desc()).limit(1000).all()
        for r in records:
            writer.writerow([r.id, r.action, r.source, r.timestamp])
            
    elif table_name == 'predictions':
        writer.writerow(['ID', 'Crop Type', 'Moisture (%)', 'Temperature (°C)', 'Humidity (%)', 'Rain Input', 'Prediction', 'Confidence (%)', 'Timestamp'])
        records = Predictions.query.order_by(Predictions.timestamp.desc()).limit(1000).all()
        for r in records:
            writer.writerow([r.id, r.crop_type, r.moisture, r.temperature, r.humidity, r.rain, r.prediction, r.confidence, r.timestamp])
            
    output.seek(0)
    return output.getvalue()

def export_excel(table_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table_name.replace('_', ' ').title()
    
    # Styles
    title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C5E3B', end_color='2C5E3B', fill_type='solid') # Forest Green
    title_fill = PatternFill(start_color='1E3F29', end_color='1E3F29', fill_type='solid') # Dark Green
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Generate Header and rows
    if table_name == 'sensor_data':
        ws.merge_cells('A1:F1')
        ws['A1'] = "SENSOR LOGS REPORT"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        
        headers = ['ID', 'Moisture (%)', 'Temperature (°C)', 'Humidity (%)', 'Rain Value', 'Timestamp']
        ws.append(headers)
        
        records = SensorData.query.order_by(SensorData.timestamp.desc()).limit(1000).all()
        for r in records:
            ws.append([r.id, r.moisture, r.temperature, r.humidity, r.rain_value, r.timestamp.strftime('%Y-%m-%d %H:%M:%S')])
            
    elif table_name == 'pump_logs':
        ws.merge_cells('A1:D1')
        ws['A1'] = "PUMP HISTORY REPORT"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        
        headers = ['ID', 'Action', 'Source', 'Timestamp']
        ws.append(headers)
        
        records = PumpLogs.query.order_by(PumpLogs.timestamp.desc()).limit(1000).all()
        for r in records:
            ws.append([r.id, r.action, r.source, r.timestamp.strftime('%Y-%m-%d %H:%M:%S')])
            
    elif table_name == 'predictions':
        ws.merge_cells('A1:I1')
        ws['A1'] = "ML PREDICTION HISTORY REPORT"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        
        headers = ['ID', 'Crop Type', 'Moisture (%)', 'Temperature (°C)', 'Humidity (%)', 'Rain Input', 'Prediction', 'Confidence (%)', 'Timestamp']
        ws.append(headers)
        
        records = Predictions.query.order_by(Predictions.timestamp.desc()).limit(1000).all()
        for r in records:
            ws.append([r.id, r.crop_type, r.moisture, r.temperature, r.humidity, r.rain, r.prediction, r.confidence, r.timestamp.strftime('%Y-%m-%d %H:%M:%S')])
            
    # Format Headers (row 2)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    # Format data cells
    for row in range(3, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col != ws.max_column:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    # Output to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def generate_pdf_report():
    """
    Generates a premium looking PDF report with ReportLab.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    story = []
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#2C5E3B'),
        spaceAfter=15,
        alignment=1 # Center
    )
    
    sub_title_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor('#555555'),
        spaceAfter=25,
        alignment=1 # Center
    )
    
    h2_style = ParagraphStyle(
        'Heading2_Green',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#1E3F29'),
        spaceBefore=12,
        spaceAfter=8
    )
    
    cell_style = ParagraphStyle(
        'GridCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11
    )
    
    cell_header_style = ParagraphStyle(
        'GridHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white
    )

    # 1. Header Section
    story.append(Paragraph("Smart Irrigation System Project Report", title_style))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Admin Executive Copy", sub_title_style))
    story.append(Spacer(1, 10))
    
    # 2. Brief summary tables
    story.append(Paragraph("System Configuration & Summary", h2_style))
    
    total_users = User.query.count()
    total_logs = SensorData.query.count()
    total_predictions = Predictions.query.count()
    total_pump_activations = PumpLogs.query.filter_by(action='Pump ON').count()
    
    summary_data = [
        [Paragraph("Metric", cell_header_style), Paragraph("Value", cell_header_style)],
        [Paragraph("Total Registered Users", cell_style), Paragraph(str(total_users), cell_style)],
        [Paragraph("Total Sensor Logs Recorded", cell_style), Paragraph(str(total_logs), cell_style)],
        [Paragraph("Total ML Predictions Made", cell_style), Paragraph(str(total_predictions), cell_style)],
        [Paragraph("Total Pump ON Activations", cell_style), Paragraph(str(total_pump_activations), cell_style)]
    ]
    
    summary_table = Table(summary_data, colWidths=[200, 150])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2C5E3B')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F2F7F4')])
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # 3. Recent logs section
    story.append(Paragraph("Recent Sensor Activity Logs (Last 15 Records)", h2_style))
    
    sensor_records = SensorData.query.order_by(SensorData.timestamp.desc()).limit(15).all()
    sensor_table_data = [
        [Paragraph("ID", cell_header_style), 
         Paragraph("Soil Moisture (%)", cell_header_style), 
         Paragraph("Temp (°C)", cell_header_style), 
         Paragraph("Humidity (%)", cell_header_style), 
         Paragraph("Rain Value", cell_header_style), 
         Paragraph("Timestamp", cell_header_style)]
    ]
    
    for r in sensor_records:
        sensor_table_data.append([
            Paragraph(str(r.id), cell_style),
            Paragraph(f"{r.moisture:.2f}%", cell_style),
            Paragraph(f"{r.temperature:.1f}°C", cell_style),
            Paragraph(f"{r.humidity:.1f}%", cell_style),
            Paragraph(f"{r.rain_value:.1f}", cell_style),
            Paragraph(r.timestamp.strftime('%Y-%m-%d %H:%M:%S'), cell_style)
        ])
        
    sensor_table = Table(sensor_table_data, colWidths=[40, 100, 70, 90, 80, 140])
    sensor_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2C5E3B')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FBF9')])
    ]))
    
    story.append(sensor_table)
    story.append(Spacer(1, 20))
    
    # 4. Recent Pump Actions
    story.append(Paragraph("Recent Pump Actions (Last 10 Records)", h2_style))
    pump_records = PumpLogs.query.order_by(PumpLogs.timestamp.desc()).limit(10).all()
    
    pump_table_data = [
        [Paragraph("ID", cell_header_style), 
         Paragraph("Action Taken", cell_header_style), 
         Paragraph("Trigger Source", cell_header_style), 
         Paragraph("Timestamp", cell_header_style)]
    ]
    
    for r in pump_records:
        pump_table_data.append([
            Paragraph(str(r.id), cell_style),
            Paragraph(r.action, cell_style),
            Paragraph(r.source, cell_style),
            Paragraph(r.timestamp.strftime('%Y-%m-%d %H:%M:%S'), cell_style)
        ])
        
    pump_table = Table(pump_table_data, colWidths=[50, 120, 120, 230])
    pump_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3F29')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FBF9')])
    ]))
    
    story.append(pump_table)
    
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
